import openpyxl as xl
from typing import Dict, List

sale_headers = ['AE PN', 'Customer', 'Customer PN', 'PO REF', 'From Stock', 'Qty', 'Price', 'Discount', 'Shipment',
                'Final Price']  # Replace with actual headers

def load_data(worksheet: xl.Worksheet, headers: List[str], data_dict: Dict) -> None:
    """Loads data from a worksheet into a dictionary.

    Args:
        worksheet: The worksheet to load data from.
        headers: The column headers for the data.
        data_dict: The dictionary to store the loaded data.
    """

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        key = row[0]
        data_dict[key] = dict(zip(headers, row[1:]))

def load_sales(worksheet: xl.Worksheet, data: Dict) -> None:
    """Loads sales data from a worksheet.

    Args:
        worksheet: The worksheet to load data from.
        data: The dictionary to store the loaded data.
    """

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        key, ae_ref = row[0], row[3]
        data.setdefault(key, {}).setdefault(ae_ref, dict(zip(sale_headers, row[1:])))

def query_sales(data: Dict, ae_pn: int) -> Dict:
    """Queries sales data by AE PN.

    Args:
        data: The dictionary containing sales data.
        ae_pn: The AE PN to query.

    Returns:
        A dictionary containing sales data for the specified AE PN,
        or an error message if the PN is not found.
    """

    ae_pn = str(ae_pn)
    if ae_pn not in data:
        return f"AE PN {ae_pn} not found"

    customer = input("Enter Customer: ")
    if not customer:
        return data[ae_pn]

    result = {f"Sale history of {ae_pn} to {customer}": {}}
    result.update({key: value for key, value in data[ae_pn].items() if value["Customer"] == customer})
    return result

def add_sale(worksheet: xl.Worksheet, data: Dict) -> None:
    """Adds a new sale to a worksheet.

    Args:
        worksheet: The worksheet to add the sale to.
        data: The dictionary containing sales data.
    """

    while True:
        ae_pn = input("Enter AE PN: ")
        if ae_pn in data:
            break
        print("AE PN not found")

    sale_data = [
        int(ae_pn),
        input("Enter Customer: "),
        input("Enter Customer PN: "),
        input("Enter PO REF: "),
        input("Enter From Stock: "),
        int(input("Enter Qty: ")),
        int(input("Enter Price: ")),
        float(input("Enter Discount: ")) / 100,
        int(input("Enter Shipment: "))
    ]

    sale_data.append(f"={sale_data[5]}*{sale_data[6]}*(1-{sale_data[7]})+{sale_data[8]}")
    worksheet.append(sale_data)

def main():
    """The main function that runs the program."""

    src = input("Please drag file here or enter file path: ")
    wb = xl.load_workbook(src)
    ps = wb["Data"]
    sd = wb["Test"]

    data_headers = ['AE PN', 'Valve', 'Actuator', 'Linkage', 'LD', 'Limit Switch', 'Positioner', 'Linkage',
                'Valve Repair Kit', 'Ormat PN']  # Replace with actual headers
    sale_headers = ['AE PN', 'Customer', 'Customer PN', 'PO REF', 'From Stock', 'Qty', 'Price', 'Discount', 'Shipment',
                'Final Price']  # Replace with actual headers

    data = {}
    load_data(ps, data_headers, data)

    sales = {}
    load_sales(sd, sales)

    while True:
        print("Welcome to the AE Sales History App")
        print("1. View Sales Data")
        print("2. Add Sale Data")
        print("3. View AE Data")
        print("4. Search for AE PN Sale Data")
        print("5. Exit")
        choice = input("Enter your choice: ")

        if choice == "1":
            print("All Sales Data:")
            print(sales)
        elif choice == "2":
            add_sale(sd, sales)
        elif choice == "3":
            print("AE Data:")
            print(data)
        elif choice == "4":
            ae_pn = input("Enter AE PN: ")
            result = query_sales(sales, ae_pn)
            print(result)
        elif choice == "5":
            break
        else:
            print("Invalid choice. Please enter a valid option.")

    wb.close()

if __name__ == "__main__":
    main()
